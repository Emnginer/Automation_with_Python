import openpyxl as xl
from openpyxl.chart import BarChart, Reference


wb = xl.load_workbook("book1.xlsx")
sheet = wb["Sheet1"]
#cell = wb["a1"] name of the column
#cell = sheet.cell(1,2) another process
#print(cell.value) add value
# print(sheet.max_row) # How many rows in this spreadsheet

for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)
    corrected_price = float(cell.value or 0 ) * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price


Values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(Values)
sheet.add_chart(chart, 'e2')

wb.save('book1.xlsx')
